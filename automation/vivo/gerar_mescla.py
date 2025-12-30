# Primeira etapa da cotação © Victor Carbajo 
# Input: diretório com as planilhas de cotações vazias, baixadas do portal Vivo.
# Output: Mescla com todas as cotações.

import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# --- Cabeçalhos esperados (A até AG), na ordem exata ---
COLUMNS = [
    "Codigo","Produto","Velocidade","UF - A","Município - A","Endereço - A","CEP - A",
    "Latitude - A","Longitude - A","UF - B","Município - B","Endereço - B","CEP - B",
    "Latitude - B","Longitude - B","Prazo","Unidade de Medida","SLA","Interface",
    "Tipo de Conector","Tipo de Proteção","Prazo de ativação",
    "Mensalidade 12 Meses líquido","Taxa de instalação líquida 12 meses",
    "Mensalidade 24 Meses líquido","Taxa de instalação líquida 24 meses",
    "Mensalidade 36 Meses líquido","Taxa de instalação líquida 36 meses",
    "Mensalidade 48 Meses líquido","Taxa de instalação líquida 48 meses",
    "Mensalidade 60 Meses líquido","Taxa de instalação líquida 60 meses","Observações"
]

# --- GUI de seleção ---
root = tk.Tk()
root.withdraw()

file_paths = filedialog.askopenfilenames(
    title="Selecione as planilhas (todas com colunas A:AG)",
    filetypes=[("Excel (*.xlsx, *.xls)", "*.xlsx *.xls"), ("All files", "*.*")]
)
if not file_paths:
    print("Nenhum arquivo selecionado.")
    raise SystemExit

# Onde salvar
output_file = filedialog.asksaveasfilename(
    title="Salvar tabela mesclada como",
    defaultextension=".xlsx",
    filetypes=[("Excel (*.xlsx)", "*.xlsx")]
)
if not output_file:
    print("Saída não escolhida.")
    raise SystemExit

# --- Leitura e normalização ---
dataframes = []
for file in file_paths:
    # Lê usando o cabeçalho da própria planilha (linha 1) e restringe às colunas A:AG
    # (sem skiprows! isso evita o deslocamento)
    try:
        df = pd.read_excel(file, header=0, usecols="A:AG")
    except Exception as e:
        messagebox.showerror("Erro ao ler arquivo", f"{os.path.basename(file)}\n{e}")
        continue

    # Padroniza nomes (tira espaços extras) e força a ordem/nomes esperados
    df.columns = [str(c).strip() for c in df.columns]
    # Se vierem nomes diferentes mas a ordem estiver correta, forçamos os nomes esperados:
    if df.shape[1] == len(COLUMNS):
        df.columns = COLUMNS
    else:
        # Tenta selecionar apenas as colunas esperadas; se faltar alguma, cria vazia
        present = [c for c in COLUMNS if c in df.columns]
        missing = [c for c in COLUMNS if c not in df.columns]
        df = df[present]
        for c in missing:
            df[c] = pd.NA
        # Reordena exatamente como COLUMNS
        df = df[COLUMNS]

    # Remove linhas totalmente vazias
    df = df.dropna(how="all", subset=COLUMNS)

    # Adiciona a coluna de origem
    df["Cotação"] = os.path.basename(file)

    # Se ainda restou algo, guarda
    if not df.empty:
        dataframes.append(df)

if not dataframes:
    messagebox.showwarning("Aviso", "Nenhuma tabela com dados encontrada.")
    raise SystemExit

# --- Concatena verticalmente (uma embaixo da outra) ---
merged_df = pd.concat(dataframes, ignore_index=True)

# --- Salva Excel (com cabeçalho único) ---
merged_df.to_excel(output_file, index=False)

# --- Coloração alternada por bloco (cada arquivo uma cor) ---
wb = load_workbook(output_file)
ws = wb.active

fill1 = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # azul claro
fill2 = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")  # azul escuro

start_row = 2  # após o cabeçalho
toggle = True
max_col = ws.max_column  # inclui a coluna "Cotação"

for df in dataframes:
    rows_count = len(df)
    fill = fill1 if toggle else fill2
    for r in range(start_row, start_row + rows_count):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).fill = fill
    start_row += rows_count
    toggle = not toggle

# (opcional) Congela o cabeçalho
ws.freeze_panes = "A2"

wb.save(output_file)

messagebox.showinfo("Tudo certo!", f"Mesclagem concluída.\nArquivo salvo em:\n{output_file}")
print(f"Mesclagem concluída! Arquivo salvo como {output_file}")