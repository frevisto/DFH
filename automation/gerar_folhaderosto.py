import os
import re
from tkinter import (
    Tk, Label, Entry, Button,
    messagebox, filedialog, ttk
)
from openpyxl import load_workbook
import pandas as pd

# ================= CONFIGURAÇÕES =================

MODELO_PATH = "../data/static/FR_Modelo.xlsx"
OUTDIR = "../outdir"

CAMPOS_CELULAS = {
    "Designação": "B1",
    "ID Contrato": "B3",
    "Cliente da DFH": "B5",
    "Cliente Final": "B7",
    "Endereço": "B9",
    "Número": "B11",
    "Bairro": "B13",
    "Cidade": "B15",
    "UF": "B17",
    "CEP": "B19",
    "Nome do Provedor": "B21",
    "Data de Ativação": "B23",
    "Data de Vencimento do Boleto": "B25",
    "Data do Reajuste": "B27",
    "Multa/Fidelidade": "B29",
    "Cálculo da Multa": "B31",
    "Preço da 1ª Mensalidade": "B33",
    "Preço da Mensalidade Atual": "B35",
    "Preço da Instalação": "B37",
    "Quanto Cobramos Mensalidade": "B39",
    "Quanto Cobramos Instalação": "B41"
}

CAMPOS_OBRIGATORIOS = [
    "Cliente da DFH",
    "Cliente Final",
    "Cidade",
    "UF",
    "Nome do Provedor"
]

CAMPOS_DATA = {
    "Data de Ativação",
    "Data de Vencimento do Boleto",
    "Data do Reajuste"
}

CAMPOS_VALOR = {
    "Multa/Fidelidade",
    "Cálculo da Multa",
    "Preço da 1ª Mensalidade",
    "Preço da Mensalidade Atual",
    "Preço da Instalação",
    "Quanto Cobramos Mensalidade",
    "Quanto Cobramos Instalação"
}

# ================= FUNÇÃO CENTRAL =================

def gerar_folha_dados(dados: dict):
    wb = load_workbook(MODELO_PATH)
    ws = wb.active

    for campo in CAMPOS_OBRIGATORIOS:
        if not dados.get(campo):
            raise ValueError(f"O campo obrigatório '{campo}' não foi preenchido.")

    for campo, celula in CAMPOS_CELULAS.items():
        valor = dados.get(campo, "").strip()
        ws[celula] = valor if valor else "-"

    os.makedirs(OUTDIR, exist_ok=True)
    nome_arquivo = f"{dados['Cliente Final']}.xlsx"
    caminho_saida = os.path.join(OUTDIR, nome_arquivo)

    wb.save(caminho_saida)

# ================= MÁSCARAS =================

def mascara_cep(entry):
    texto = re.sub(r"\D", "", entry.get())[:8]
    if len(texto) > 5:
        texto = f"{texto[:5]}-{texto[5:]}"
    entry.delete(0, "end")
    entry.insert(0, texto)

def mascara_data(entry):
    texto = re.sub(r"\D", "", entry.get())[:8]
    if len(texto) >= 5:
        texto = f"{texto[:2]}/{texto[2:4]}/{texto[4:]}"
    elif len(texto) >= 3:
        texto = f"{texto[:2]}/{texto[2:]}"
    entry.delete(0, "end")
    entry.insert(0, texto)

def mascara_valor(entry):
    texto = re.sub(r"[^\d]", "", entry.get())
    if texto:
        valor = int(texto)
        formatado = f"R$ {valor/100:,.2f}"
        formatado = formatado.replace(",", "X").replace(".", ",").replace("X", ".")
        entry.delete(0, "end")
        entry.insert(0, formatado)

# ================= GUI MANUAL =================

def gui_manual():
    root = Tk()
    root.title("Gerar Folha de Rosto - Manual")
    root.geometry("600x800")
    root.resizable(True, True)

    style = ttk.Style(root)
    style.theme_use("clam")

    style.configure("Title.TLabel", font=("Segoe UI", 13, "bold"))
    style.configure("Form.TLabel", font=("Segoe UI", 10))
    style.configure("Action.TButton", font=("Segoe UI", 10), padding=6)

    # ===== Container principal =====
    main = ttk.Frame(root, padding=20)
    main.pack(expand=True, fill="both")

    entradas = {}
    linha = 1

    # ===== Campos =====
    for campo in CAMPOS_CELULAS:
        ttk.Label(
            main,
            text=campo,
            style="Form.TLabel"
        ).grid(row=linha, column=0, sticky="e", padx=10, pady=6)

        entry = ttk.Entry(main, width=42)
        entry.grid(row=linha, column=1, padx=10, pady=6, sticky="w")

        if campo == "CEP":
            entry.bind("<KeyRelease>", lambda e, ent=entry: mascara_cep(ent))
        elif campo in CAMPOS_DATA:
            entry.bind("<KeyRelease>", lambda e, ent=entry: mascara_data(ent))
        elif campo in CAMPOS_VALOR:
            entry.bind("<KeyRelease>", lambda e, ent=entry: mascara_valor(ent))

        entradas[campo] = entry
        linha += 1

    # ===== Ações =====
    def gerar():
        try:
            dados = {campo: ent.get() for campo, ent in entradas.items()}
            gerar_folha_dados(dados)
            messagebox.showinfo("Sucesso", "Folha de rosto gerada com sucesso na pasta ./outdir ")
        except Exception as e:
            messagebox.showerror("Erro", str(e))

    botoes = ttk.Frame(main, padding=(0, 20))
    botoes.grid(row=linha, column=0, columnspan=2)

    ttk.Button(
        botoes,
        text="Gerar Folha",
        style="Action.TButton",
        command=gerar
    ).grid(row=0, column=0, padx=10)

    ttk.Button(
        botoes,
        text="Voltar ao Menu",
        style="Action.TButton",
        command=lambda: [root.destroy(), menu_inicial()]
    ).grid(row=0, column=1, padx=10)

    root.mainloop()

# ================= GUI LOTE =================

def gui_lote():
    try:
        arquivo = filedialog.askopenfilename(
            title="Selecione a planilha de lote",
            filetypes=[("Excel", "*.xlsx")]
        )

        if not arquivo:
            return

        df = pd.read_excel(arquivo)

        # Verificação de planilha vazia
        if df.empty:
            raise ValueError(
                "A planilha selecionada não contém dados.\n\n"
                "Preencha ao menos uma linha antes de continuar."
            )

        # Validação de cabeçalho
        colunas_esperadas = set(CAMPOS_CELULAS.keys())
        colunas_encontradas = set(df.columns)

        colunas_faltantes = colunas_esperadas - colunas_encontradas
        if colunas_faltantes:
            raise ValueError(
                "A planilha está sem as seguintes colunas obrigatórias:\n\n"
                + "\n".join(sorted(colunas_faltantes))
            )

        for _, row in df.iterrows():
            dados = {
                campo: str(row[campo]).strip()
                if not pd.isna(row[campo]) else ""
                for campo in CAMPOS_CELULAS
            }
            gerar_folha_dados(dados)

        messagebox.showinfo(
            "Sucesso",
            f"Geração em lote concluída.\n\nTotal de folhas geradas: {len(df)}"
        )

    except Exception as e:
        messagebox.showerror("Erro", str(e))


#=============INSTRUÇÕES============

def mostrar_instrucoes_lote():
    janela = Tk()
    janela.title("Instruções – Geração em Lote")
    janela.geometry("560x510")
    janela.resizable(True, True)

    style = ttk.Style(janela)
    style.theme_use("clam")

    style.configure("Title.TLabel", font=("Segoe UI", 14, "bold"))
    style.configure("Subtitle.TLabel", font=("Segoe UI", 10))
    style.configure("Section.TLabel", font=("Segoe UI", 11, "bold"))
    style.configure("Text.TLabel", font=("Segoe UI", 10))
    style.configure("Action.TButton", font=("Segoe UI", 10), padding=8)

    # ===== Container =====
    container = ttk.Frame(janela, padding=20)
    container.pack(expand=True, fill="both")

    # ===== Título =====
    ttk.Label(
        container,
        text="Geração de Folhas de Rosto em Lote",
        style="Title.TLabel"
    ).pack(anchor="w", pady=(0, 5))

    ttk.Label(
        container,
        text="Utilize a planilha modelo para gerar múltiplos arquivos automaticamente.",
        style="Subtitle.TLabel"
    ).pack(anchor="w", pady=(0, 15))

    # ===== Seção: Modelo =====
    ttk.Label(
        container,
        text="📄 Planilha Modelo",
        style="Section.TLabel"
    ).pack(anchor="w", pady=(0, 4))

    ttk.Label(
        container,
        text="O modelo está disponível em:\n./data/static/planilha_Lote.xlsx",
        style="Text.TLabel"
    ).pack(anchor="w", pady=(0, 15))

    # ===== Seção: Instruções =====
    ttk.Label(
        container,
        text="📝 Instruções Importantes",
        style="Section.TLabel"
    ).pack(anchor="w", pady=(0, 6))

    instrucoes = (
        "• Não altere os nomes das colunas, no cabeçalho.\n"
        "• Cada linha da planilha gera uma folha de rosto.\n"
        "• Campos opcionais podem permanecer em branco.\n"
        "• O nome do arquivo será definido pelo campo 'Cliente Final'."
    )

    ttk.Label(
        container,
        text=instrucoes,
        style="Text.TLabel",
        justify="left"
    ).pack(anchor="w", pady=(0, 15))

    # ===== Seção: Obrigatórios =====
    ttk.Label(
        container,
        text="⚠️ Campos Obrigatórios",
        style="Section.TLabel"
    ).pack(anchor="w", pady=(0, 6))

    obrigatorios = (
        "• Cliente da DFH\n"
        "• Cliente Final\n"
        "• Cidade\n"
        "• UF\n"
        "• Nome do Provedor"
    )

    ttk.Label(
        container,
        text=obrigatorios,
        style="Text.TLabel",
        justify="left"
    ).pack(anchor="w", pady=(0, 20))

    # ===== Botões =====
    botoes = ttk.Frame(container)
    botoes.pack(pady=10)

    ttk.Button(
        botoes,
        text="Continuar para Selecionar a Planilha",
        style="Action.TButton",
        command=lambda: [janela.destroy(), gui_lote()]
    ).pack(side="left", padx=10)

    ttk.Button(
        botoes,
        text="Voltar ao Menu",
        style="Action.TButton",
        command=lambda: [janela.destroy(), menu_inicial()]
    ).pack(side="left", padx=10)


    janela.mainloop()

#========== AJUDA ===========
def mostrar_ajuda():
    janela = Tk()
    janela.title("Ajuda – Campos Obrigatórios")
    janela.geometry("420x360")
    janela.resizable(False, False)

    style = ttk.Style(janela)
    style.theme_use("clam")

    style.configure("Title.TLabel", font=("Segoe UI", 13, "bold"))
    style.configure("Text.TLabel", font=("Segoe UI", 10))
    style.configure("Action.TButton", font=("Segoe UI", 10), padding=6)

    container = ttk.Frame(janela, padding=20)
    container.pack(expand=True, fill="both")

    ttk.Label(
        container,
        text="Campos Obrigatórios",
        style="Title.TLabel"
    ).pack(anchor="w", pady=(0, 10))

    texto = (
        "Os seguintes campos são obrigatórios tanto\n"
        "na geração manual quanto na geração em lote:\n\n"
        "• Cliente da DFH\n"
        "• Cliente Final\n"
        "• Cidade\n"
        "• UF\n"
        "• Nome do Provedor\n\n"
        "Caso algum desses campos não seja preenchido,\n"
        "a folha de rosto não será gerada."
    )

    ttk.Label(
        container,
        text=texto,
        style="Text.TLabel",
        justify="left"
    ).pack(anchor="w", pady=(0, 20))

    ttk.Button(
        container,
        text="Fechar",
        style="Action.TButton",
        command=janela.destroy
    ).pack()

    janela.mainloop()


# ================= MENU INICIAL =================

def menu_inicial():
    root = Tk()
    root.title("Gerador de Folhas de Rosto")
    root.geometry("420x320")
    root.resizable(False, False)

    style = ttk.Style(root)
    style.theme_use("clam")

    style.configure(
        "Title.TLabel",
        font=("Segoe UI", 14, "bold")
    )

    style.configure(
        "Action.TButton",
        font=("Segoe UI", 11),
        padding=10
    )

    container = ttk.Frame(root, padding=20)
    container.pack(expand=True, fill="both")

    ttk.Label(
        container,
        text="Gerador de Folhas de Rosto",
        style="Title.TLabel"
    ).pack(pady=(10, 5))

    ttk.Label(
        container,
        text="Selecione o modo de operação",
        font=("Segoe UI", 10)
    ).pack(pady=(0, 20))

    ttk.Button(
        container,
        text="📄 Gerar Manualmente",
        style="Action.TButton",
        command=lambda: [root.destroy(), gui_manual()]
    ).pack(fill="x", pady=8)

    ttk.Button(
        container,
        text="📊 Gerar em Lote",
        style="Action.TButton",
        command=lambda: [root.destroy(), mostrar_instrucoes_lote()]
    ).pack(fill="x", pady=8)

    ttk.Button(
        container,
        text="ℹ️ Ajuda – Campos Obrigatórios",
        style="Action.TButton",
        command=mostrar_ajuda
    ).pack(fill="x", pady=6)


    root.mainloop()

# ================= START =================

if __name__ == "__main__":
    menu_inicial()
